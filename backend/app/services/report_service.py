import os
import io
from datetime import datetime, timezone
from pathlib import Path
from sqlalchemy.orm import Session
from app.core.config import settings
from app.models.report import Report, ReportFormat, ReportStatus
from app.models.kpi import KPIRecord
from app.models.institution import Institution
from app.ai.chatbot import generate_report_narrative


_FORMULA_INJECTION_CHARS = frozenset("=+-@\t\r")


def _xlsx_safe(value):
    """Prevent CSV/Excel formula injection by neutralising formula-starting characters."""
    if isinstance(value, str) and value and value[0] in _FORMULA_INJECTION_CHARS:
        return " " + value
    return value


DOMAIN_LABELS = {
    "academic": "Académique",
    "finance": "Finance",
    "hr": "Ressources Humaines",
    "esg": "ESG / RSE",
    "insertion": "Insertion Professionnelle",
    "research": "Recherche",
    "infrastructure": "Infrastructure",
    "partnerships": "Partenariats",
    "student_life": "Vie Estudiantine",
}

INDICATOR_LABELS = {
    "success_rate": "Taux de réussite",
    "attendance_rate": "Taux de présence",
    "dropout_rate": "Taux d'abandon",
    "repetition_rate": "Taux de redoublement",
    "budget_allocated": "Budget alloué",
    "budget_consumed": "Budget consommé",
    "budget_execution_rate": "Taux d'exécution budgétaire",
    "cost_per_student": "Coût par étudiant",
    "teaching_headcount": "Effectif enseignant",
    "admin_headcount": "Effectif administratif",
    "absenteeism_rate": "Taux d'absentéisme",
    "training_hours": "Heures de formation",
    "energy_consumption_kwh": "Consommation énergétique (kWh)",
    "carbon_footprint_ton": "Empreinte carbone (tonnes)",
    "recycling_rate": "Taux de recyclage",
    "employability_rate": "Taux d'employabilité",
    "national_convention_rate": "Taux de convention nationale",
    "international_convention_rate": "Taux de convention internationale",
    "insertion_delay_months": "Délai d'insertion (mois)",
    "publications_count": "Nombre de publications",
    "active_projects": "Projets actifs",
    "funding_tnd": "Financements (TND)",
}


def _get_kpi_rows(db: Session, institution_id: int, period_label: str) -> list[dict]:
    records = (
        db.query(KPIRecord)
        .filter(
            KPIRecord.institution_id == institution_id,
            KPIRecord.period_label == period_label,
        )
        .order_by(KPIRecord.domain, KPIRecord.indicator_key)
        .all()
    )
    return [
        {
            "domain": r.domain,
            "domain_label": DOMAIN_LABELS.get(r.domain, r.domain),
            "indicator_key": r.indicator_key,
            "indicator_label": INDICATOR_LABELS.get(r.indicator_key, r.indicator_key),
            "value": r.value,
            "unit": r.unit,
        }
        for r in records
    ]


def generate_pdf_report(db: Session, report: Report) -> str:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    filename = f"report_{report.institution_id}_{report.period_label}_{report.id}.pdf"
    filepath = os.path.join(settings.REPORTS_DIR, filename)

    institution = db.query(Institution).filter(Institution.id == report.institution_id).first()
    inst_name = institution.name if institution else "Établissement"

    kpi_rows = _get_kpi_rows(db, report.institution_id, report.period_label)

    kpi_summary_str = "\n".join(
        f"- {r['indicator_label']}: {r['value']} {r['unit']}" for r in kpi_rows
    )
    narrative = generate_report_narrative(kpi_summary_str, report.period_label, inst_name)

    doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    UCAR_BLUE = colors.HexColor("#1B3A6B")
    UCAR_GOLD = colors.HexColor("#D4A017")

    title_style = ParagraphStyle("title", parent=styles["Title"], textColor=UCAR_BLUE, fontSize=18, spaceAfter=6)
    subtitle_style = ParagraphStyle("subtitle", parent=styles["Normal"], textColor=UCAR_GOLD, fontSize=12, spaceAfter=4)
    section_style = ParagraphStyle("section", parent=styles["Heading2"], textColor=UCAR_BLUE, fontSize=13, spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle("body", parent=styles["Normal"], fontSize=10, leading=14)

    story = []

    story.append(Paragraph("Université de Carthage — UCAR", subtitle_style))
    story.append(Paragraph(f"Rapport Institutionnel — {inst_name}", title_style))
    story.append(Paragraph(f"Période : {report.period_label}", body_style))
    story.append(Paragraph(f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}", body_style))
    story.append(HRFlowable(width="100%", thickness=2, color=UCAR_GOLD, spaceAfter=12))

    story.append(Paragraph("Résumé Exécutif", section_style))
    for para in narrative.split("\n\n"):
        if para.strip():
            story.append(Paragraph(para.strip(), body_style))
            story.append(Spacer(1, 6))

    story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey, spaceBefore=12, spaceAfter=12))

    story.append(Paragraph("Indicateurs de Performance (KPIs)", section_style))

    grouped: dict[str, list] = {}
    for row in kpi_rows:
        grouped.setdefault(row["domain_label"], []).append(row)

    for domain_label, rows in grouped.items():
        story.append(Paragraph(domain_label, ParagraphStyle("dom", parent=styles["Heading3"], textColor=UCAR_BLUE, fontSize=11, spaceBefore=8)))

        table_data = [["Indicateur", "Valeur", "Unité"]]
        for row in rows:
            table_data.append([_xlsx_safe(row["indicator_label"]), str(row["value"]), _xlsx_safe(row["unit"])])

        t = Table(table_data, colWidths=[10*cm, 4*cm, 3*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), UCAR_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 8))

    doc.build(story)
    return filepath


def generate_excel_report(db: Session, report: Report) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(settings.REPORTS_DIR, exist_ok=True)
    filename = f"report_{report.institution_id}_{report.period_label}_{report.id}.xlsx"
    filepath = os.path.join(settings.REPORTS_DIR, filename)

    institution = db.query(Institution).filter(Institution.id == report.institution_id).first()
    inst_name = institution.name if institution else "Établissement"

    kpi_rows = _get_kpi_rows(db, report.institution_id, report.period_label)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résumé"

    BLUE = "1B3A6B"
    GOLD = "D4A017"
    LIGHT = "F5F7FA"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor=BLUE)
    gold_font = Font(name="Calibri", bold=True, color=GOLD, size=11)
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:C1")
    ws["A1"] = f"Rapport Institutionnel — {inst_name}"
    ws["A1"].font = Font(name="Calibri", bold=True, color=BLUE, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:C2")
    ws["A2"] = f"Période : {report.period_label} | Généré le {datetime.now().strftime('%d/%m/%Y')}"
    ws["A2"].font = gold_font
    ws["A2"].alignment = center

    ws.append([])

    ws.append(["Indicateur", "Valeur", "Unité"])
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for i, row in enumerate(kpi_rows, start=5):
        ws.append([_xlsx_safe(row["indicator_label"]), row["value"], _xlsx_safe(row["unit"])])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = PatternFill("solid", fgColor=LIGHT)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    for domain_label, rows in _group_by_domain(kpi_rows).items():
        sheet = wb.create_sheet(title=domain_label[:31])
        sheet.append(["Indicateur", "Valeur", "Unité"])
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for row in rows:
            sheet.append([_xlsx_safe(row["indicator_label"]), row["value"], _xlsx_safe(row["unit"])])
        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 15
        sheet.column_dimensions["C"].width = 12

    wb.save(filepath)
    return filepath


def _group_by_domain(rows: list[dict]) -> dict[str, list]:
    grouped: dict[str, list] = {}
    for row in rows:
        grouped.setdefault(row["domain_label"], []).append(row)
    return grouped


def process_report(db: Session, report_id: int) -> Report:
    report = db.query(Report).filter(Report.id == report_id).first()
    if not report:
        raise ValueError(f"Rapport {report_id} introuvable")

    report.status = ReportStatus.generating
    db.commit()

    try:
        if report.format == ReportFormat.pdf:
            filepath = generate_pdf_report(db, report)
        else:
            filepath = generate_excel_report(db, report)

        report.file_path = filepath
        report.status = ReportStatus.ready
        report.generated_at = datetime.now(timezone.utc)
    except Exception as e:
        report.status = ReportStatus.failed
        raise e
    finally:
        db.commit()

    return report
